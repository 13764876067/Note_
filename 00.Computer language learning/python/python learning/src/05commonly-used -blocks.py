#coding:utf-8

'''

1.python 中的常用模块部分

标准库和第三方包
    安装和使用
    第三方包
    辅助工具isort,pylint,pylama,autopep8

搭建模块
    __init__.py,__main__.py,.py文件
    包的结构
    包编写和发布

IO流,上下文管理器
    操作txt,csv,excel文件类型

异常处理
    try,except
    自定义异常类型

搭建系统:语音合成系统
    SDK,API
    结合已学过的知识

爬虫:网络爬虫
    网络爬虫第三方包
    网络爬虫一般步骤

'''

'''

1.标准库和第三方包
    import用法,(引用注意引起名称冲突)
    import module
    import module as new_name 
    from muodule import function
    from muodule import function as new_name
    form module import *
    
    查找第三方包来源
    https://pypi.org/
    安装第三方包 pip pip3
    pip install package_name
    pypi:国内镜像安装 -i
    
    nump:矩阵计算
    pandas:数据集处理
    requests:http包
    
    辅助工具:
    isort:import排序工具
    pylint,pylama:代码静态检查工具
    autopep8:自动化格式工具
    
2.模块和包的结构与发布
    module.py文件
    #!/usr/bin/env/ python3    #!/usr/bin/python3   #解释器声明 必须文件第一行,不能有空行 Linux系统用
    # -*- coding: utf-8 -*-     #文件编码表明
    python file.py
    
    在当前路径:
    import module
    moudle.class,module.method()
    
    if __name__ == "__main__":去掉测试调试的代码
    判断语句 __name__ 是否是main方法
    dir(module)
    
    不在当前路径:
    sys.path:python搜索路径
    sys.path.append(路径)
    
    包:有一定的层次的目录结构
    .py:源文件
    __init__.py:可以进行初始化操作
    __main__.py:main()方法
    
    ├──pkg
    ├── __init__.py
    ├── __main__.py
    如果你希望 python 将一个文件夹作为 Package 对待，那么这个文件夹中必须包含一个名为 __init__.py 的文件，即使它是空的。
    如果你需要 python 将一个文件夹作为 Package 执行，那么这个文件夹中必须包含一个名为 __main__.py 的文件.
    
    包结构:
    A 
      __init__.py
        pycache__
            ...
      a.py
      b.py
    B 
      __init__.py
        pycache__
            ...
      c.py
      d.py
    __init__.py
      pycache__
         ...
     
    from . import apython
    from ..B import brust
    
    发布到 pypi中作为第三方包
        steup.py:配置
        README.md:包介绍
        LICENSE:说明
        .py
        __init__.py
    工具:
        steuptools well
        twine
       
3.文件的读写:IO流
    import os
    os.getcwd()
    
    txt文件:
    f = open() :创建文件 打开模式 r w a b + x
    f.write()
    f.read()
    f.close()
    
    上下文管理器:
    with open("a.txt",'a') as f:
           f.write()
    context manager:有两个__enter__和__exit__两个方法,分别用于处理资源管理器的初始化和资源的清理方法       
                 
    for line in f:
        print(line,end='')
    f.seek()       
    :迭代器(指针操作)
    
    csv文件:
    import csv
    with open('a.csv','w') as f
        writer = csv.writer(f)
        writer.writer(data)
    
    f = open()
    reader = csv.reader(f)
    for row inreader:
        print(row)
 
    excel文件:
    from openpyxl import Workbok
    
    操作sheet
    wb = Workbook()
    ws = wb.active
    ws.title
    ws.title = 'python'
    ws2 = wb.create_sheet('java')
    ws2.title
    wb.sheetnames
    
    写入sheet
    ws['E1'] = 111
    ws.cell(row=2,column=2,value=222)
    wb.save()
  
4.异常处理: 
    1 / 0 , a
    try:捕获
    except:抛出异常
    except(ZeroDivisionError,ValueError)
    raise:可以用来抛出一个异常
    finally:最后一定执行
    
    sys.exc_info(),获取异常信息,trackback.extract_tb():获取堆栈信息
    
    自定义异常类型:
    异常对象
    
    异常类继承结构BaseException,Exception
    
5.搭建系统:语音合成系统
    SDK:软件工具开发包
    API:应用程序接口
   
    语音合成:百度SDK,API
    from aip import AipSpeech
    """ 你的 APPID AK SK """
    APP_ID = '你的 App ID'
    API_KEY = '你的 Api Key'
    SECRET_KEY = '你的 Secret Key'
    client = AipSpeech(APP_ID, API_KEY, SECRET_KEY)
    
    client与sever的连接访问数据
    向百度服务器提交信息
    result  = client.synthesis('你好百度', 'zh', 1, {'vol': 5,})
    识别正确返回语音二进制 错误则返回dict 参照下面错误码
    if not isinstance(result, dict):
        with open('auido.mp3', 'wb') as f:
            f.write(result)
            
6.网络爬虫:
    socket:pyzmq
    web:requests,lxml,re,json
    
    import requests
    from lxml import html
    url = "http://itdiffer.com"
    page = requests.get(url).content.decode('utf-8')
    sel = html.fromstring(page)
    title=sel.xpath('//article/h2/a/text()')
    #print(page)
    print(title)
    
    1.导入相关第三方包:requests,lxml,re,json
    2.网址url,解析规则regexp,保存对象csv及相关变量声明
    3.request请求网页获取html代码
    3.利用工具包或者regexp,json等工具解析html代码获取相关的数据
    4.遍历格式化数据输出或者保存为文档
    
'''

print('---------------------------------------------标准库第三方包-----------------------------------------------------')

import math
from math import pow
print(math.pow(2,3))
print(pow(2,3))
import numpy as np
import pandas as pd
import requests as req # http包

#创建 moudle.py
class Fibs:
    def __init__(self,max):
        self.a = 0
        self.b = 1
        self.max = max

    def __iter__(self):
        return self
    def __next__(self):
        fib = self.a
        if fib > self.max:
            raise StopIteration
        self.a,self.b = self.b,self.a + self.b
        return fib

def y_yield(n):
    print('strat')
    while n > 0 :
        print("before yield")
        yield n
        n -= 1
        print("after yield")
fb = Fibs(10000)
print([fb.__next__() for i in range(10)])

# import module
# dir(moudle)
# moudle.y_yield(n)
# if __name__ == "__main__"

print('---------------------------------------------标准库第三方包-----------------------------------------------------')

print('---------------------------------------------包的结构与发布-----------------------------------------------------')

#import module
#moudle.class ,module.method()
#if __name__ == "__main__": 去掉测试调试的代码
#    sys.path:python搜索路径
#    sys.path.append(路径)

'''
    发布到 pypi中作为第三方包
        steup.py:配置
        README.md:包介绍
        LICENSE:说明
        .py
        __init__.py
    工具:
        steuptools well
        twine
'''

print('---------------------------------------------包的结构与发布-----------------------------------------------------')


print('---------------------------------------------文件操作之读写-----------------------------------------------------')

#上下文管理器
class my_open:
    def __init__(self,path, suppress=False):
        self.path = path
        self.suppress = suppress
    def read(self):
        return self.f.read( )
    def fail(self):
        print('try raise a error')
        raise ValueError
    def __enter__(self):
        '''
        可以返回self或者其他什么东西
        在with语法中，如果有as,那么这里返回的内容会被绑定到as后面的变量.上面
        :return:
        '''
        print('starting with block')
        self.f = open(self.path)
        return self
    def __exit__(self,exc_type, exc_value, exc_tb):
        '''
        如果在context manager运行的过程中出现了异常，那么异常会被传入exit作为判断的依据
        exit返回False则认为context manager失败了，这些异常会被直接重新抛出
        但是如果exit计里面处理了这些异常，而且返向了-个True.那么这些异常将不会再向外传递
        '''
        self.f.close( )
        if exc_type is None:
            print('exited normally\n' )
        else:
            print('raise an exception!' + str(exc_type))
        return self.suppress
try:
    with my_open('test.txt') as f:
        f.fail()
except ValueError:
    print('ValueError raised')

try:
    with my_open('test.txt',suppress=True) as f:
        f.fail()
except ValueError:
    print('ValueError raised')

#IO流
import os
print(os.getcwd())
#f = open() :创建文件 打开模式 r w a b + x
#f.write()
#f.writelines()
#f.read()
#f.close()

#with open("a.txt",a) as f:
#   f.write()

#for line in f:
#   print(line,end='')
#f.seek()

#import csv

#with open('a.csv', 'w') as f
#    writer = csv.writer(f)
#    writer.writer(data)

#f = open()
#reader = csv.reader(f)
#for row inreader:
#    print(row)

#from openpyxl import Workbook
#wb = Workbook()
#ws = wb.active
#ws.title
#ws.title = 'python'
#ws2 = wb.create_sheet('java')
#ws2.title
#wb.sheetnames

#ws['E1'] = 111
#ws.cell(row=2,column=2,value=222)
#wb.save()

print('---------------------------------------------文件操作之读写-----------------------------------------------------')


print('---------------------------------------------异 常   处 理-----------------------------------------------------')

class Calculator:
    is_raise = False
    def calc(self,express):
        try:
            return eval(express)
        except ZeroDivisionError:
            if self.is_raise:
                return "zero can no be division"
            else:
                raise

c = Calculator()
print(c.calc('8/1'))

'''
while True:
    try:
        a = float(input())
        b = float(input())
        r = a / b
        print(a,b,r)
        break
    except ZeroDivisionError:
        print("zero")
    except ValueError:
        print('number')
    except:
        break
'''

print('---------------------------------------------异 常   处 理-----------------------------------------------------')

print('---------------------------------------------语音合成  sdk-----------------------------------------------------')

#from aip import AipSpeech

#""" 你的 APPID AK SK """
#APP_ID = '你的 App ID'
#API_KEY = '你的 Api Key'
#SECRET_KEY = '你的 Secret Key'

#client = AipSpeech(APP_ID, API_KEY, SECRET_KEY)

#client与sever的连接访问数据
#向百度服务器提交信息
#result = client.synthesis('你好百度', 'zh', 1, {'vol': 5, })
#识别正确返回语音二进制
#错误则返回dict
#参照下面错误码
#if not isinstance(result, dict):
#    with open('auido.mp3', 'wb') as f:
#        f.write(result)

print('---------------------------------------------语音合成  sdk-----------------------------------------------------')

print('---------------------------------------------爬        虫-----------------------------------------------------')

import requests
from lxml import html
url = "http://itdiffer.com"
page = requests.get(url).content.decode('utf-8')
sel = html.fromstring(page)
title=sel.xpath('//article/h2/a/text()')
#print(page)
print(title)

import re
import json
import pandas as pd
import requests
from datetime import datetime

url='https://ncow.dxy.cn/ncovh5/view/pneumonia'
page = requests.get(url).content.decode('utf-8')
regexp = "<script id=\"getListByCountryTypeServicel\">([^<]+)"
res = re.findall(regexp,page)
data = res[0][44:-11]
dicts = json.loads(data)

for row in dicts:
    for key in row:
        if key in ["createTime","modifyTime"]:
            row[key] = datetime.fromtimestamp(row[key]/1000).strptime("%Y-%m-%d %H:%M:%S")
        print(key,":",row[key],end=" ")
    print("\n")

#df = pd.DataFrame(dicts)
#df.to_csv(" ",mode="a")

print('---------------------------------------------爬        虫-----------------------------------------------------')